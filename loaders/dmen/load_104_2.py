from db_oracle.connect import get_connection
import config as cfg
from openpyxl import load_workbook
import datetime
import os.path


def load_103_2():
    s_now = datetime.datetime.now()
    # Нормируем путь к файлу по слэшам
    f_path = cfg.file_path
    f_name = cfg.file_104_2
    path = f_path + '\\' + f_name
    file_path = os.path.normpath(path)
    print("Работаем с файлом: " + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return 

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_path)

    wb = load_workbook(file_path)
    print("Книга загружена: " + path)
    sheet = wb.active

    print("Подключаем БД")

    con = get_connection()
    cursor = con.cursor()
    # Создадим новое задание
#    cursor.execute('truncate table mrr_overpay_new_3')
    print("Очистили таблицу БД mrr_overpay_new_3")
    print("Загружаем со строки: " + str(cfg.first_row_103_2))
    print("Всего в документе строк: " + str(sheet.max_row))

    count_rows = 0
    for i in range(cfg.first_row_103_2, sheet.max_row+1):
           if not sheet.cell(row=i, column=1).value:
              break	        
#           print('Запись: ' + str(i))
           cmd = "insert into  MRR_OVERPAY_NEW_3 q " \
             "(n, rfbn, msolid, rfpm, fio, birthdate, iin, snumb, date_address, appoint_date, ist, sum, priz, insp_spec, insp_ruk, insp_dir, " \
             "date_spec, date_ruk, date_dir, stopdate, reason, no_gk, mnth, is_stop )" \
             "values ( " 
           for x in range(1,23):
#              print('Колонка ' + str(x) + ' : ' + sheet.cell(row=i, column=x).value )
              if isinstance(sheet.cell(row=i, column=x).value, str):
                 cmd = cmd + '\'' + sheet.cell(row=i, column=x).value + "', "
              else:
                 cmd = cmd + sheet.cell(row=i, column=x).value + ", "
           cmd = cmd + " to_date('" + cfg.month + "','dd.mm.yyyy'),'sus')"
           count_rows=count_rows+1
#           print(str(i)+' : ' + cmd)
           cursor.execute(cmd)

    con.commit()
    con.close()
    now = datetime.datetime.now()
    print("Загружено записей: " + str(count_rows) + ' : ' + now.strftime("%d-%m-%Y %H:%M:%S"))
    return


if __name__ == "__main__":
    print("APP TestingAdmin starting")
    load_103_2()
