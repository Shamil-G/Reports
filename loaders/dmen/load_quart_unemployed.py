from db_oracle.connect import get_connection
import cx_Oracle
import configs.config_unemployed as cfg
from openpyxl import load_workbook
import datetime
import os.path
import xlsxwriter


class ResultF(object):
    def __init__(self, id, region, cz_code, cz_name, iin, lastname, firstname, middlename, birthday,
                 reg_date, status_code, status_in_cz, member, kzd, su, so, su24, so24, sd, is0703, pm, nsu, ksu_new,
                 size_new):
        self.id = id
        self.region = region
        self.cz_code = cz_code
        self.cz_name = cz_name
        self.iin = iin
        self.lastname = lastname
        self.firstname = firstname
        self.middlename = middlename
        self.birthday = birthday
        self.reg_date = reg_date
        self.status_code = status_code
        self.status_in_cz = status_in_cz
        self.member = member
        self.kzd = kzd
        self.su = su
        self.so = so
        self.su24 = su24
        self.so24 = so24
        self.sd = sd
        self.is0703 = is0703
        self.pm = pm
        self.nsu = nsu
        self.ksu_new = ksu_new
        self.size_new = size_new


def create_table_name():
    table_name = "TMP_DMEN_Q" + str(cfg.quart) + "_21"
    return table_name


def create_table(table_name):
    con = get_connection()
    cursor = con.cursor()
    try:
        print("Удаляем таблицу " + table_name)
        cursor.execute("drop table " + table_name)
        print("Удалена таблица " + table_name)
    except cx_Oracle.Error:
        print("Error import cx_Oracle :", cx_Oracle.DataError)

    cmd = "create table " + table_name + " " \
        "( mn nvarchar2(12), id NUMBER(8), region VARCHAR2(4), cz_code VARCHAR2(9), " \
        "cz_name VARCHAR2(128), iin VARCHAR2(12), lastname NVARCHAR2(30), " \
        "firstname NVARCHAR2(30), middlename VARCHAR2(30), birthday DATE, " \
        "reg_date DATE, status_code VARCHAR2(6), " \
        "status_in_cz VARCHAR2(16), member VARCHAR2(8), kzd number(4,2), sicid number(12), mnth date, " \
        "common_SU number(3), count_so_24 number(2), " \
        "smd24 number(19,2), count_acc_month number, " \
        "svpr_su number(2), ksu number(5,3), " \
        "pm NUMBER(16, 2) )"
    cursor.execute(cmd)
    print("Создана таблица " + table_name)
    con.commit()
    cursor.close()
    con.close()


def fill_sicid_mnth(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Начато обновление SIСID. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'begin ' \
          'for cur in (select t.rowid, p.sicid, t.iin from ' + table_name + ' t, person p '\
          ' where t.sicid=0 '\
          ' and p.rn=t.iin) '\
          '    loop '\
          ' update ' +\
          table_name + ' '\
          't2 '\
          'set t2.sicid = cur.sicid '\
          'where t2.iin = cur.iin; '\
          'end '\
          'loop; '\
          'commit; '\
          'end;'
    print('CMD: ' + cmd)
    cursor.execute(cmd)
    print("Обновлены SIСID. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cursor.execute("UPDATE " + table_name + " m SET m.mnth = trunc(reg_date,'MONTH')")
    print("Обновлены MNTH. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    con.commit()
    cursor.close()
    con.close()


def create_index(t_name):
    cmd1 = 'create index XN_' + t_name + '_IIN on ' + t_name + '(IIN)'
    cmd2 = 'create index XN_' + t_name + '_SICID on ' + t_name + '(SICID)'
    con = get_connection()
    cursor = con.cursor()
    cursor.execute(cmd1)
    cursor.execute(cmd2)
    cursor.close()
    con.close()


def load_quart_unemployed(table_name):
    s_now = datetime.datetime.now()
    # Нормируем путь к файлу по слэшам
    f_path = cfg.REPORTS_PATH
    f_name = cfg.load_file
    path = f_path + '\\' + f_name
    file_path = os.path.normpath(path)
    print("Работаем с файлом: " + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return

    print("Загрузка стартовала: " + s_now.strftime("%d-%m-%Y %H:%M:%S") + ' : ' + file_path)

    wb = load_workbook(file_path)
    print("Книга загружена: " + path)
    # sheet = wb.active
    # Создадим новое задание

    con = get_connection()
    cursor = con.cursor()

    # print('SHEET name :' + wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print('Загружается книга: ' + sheet_name)
        print("Всего в книге строк: " + str(sheet.max_row))
        count_rows = 0
        for i in range(cfg.quart_first_row, sheet.max_row+1):
            if not sheet.cell(row=i, column=1).value:
                break
    #           print('Запись: ' + str(i))
            cmd = "insert into " + table_name + " " \
               "( mn, id, region, cz_code, " \
               "cz_name, iin, lastname, firstname, middlename, " \
               "reg_date, status_code, " \
               "status_in_cz, member, kzd, sicid) " \
                "values ( \'" + str(sheet_name) + "\', "
            for x in range(1, 14):
                if isinstance(sheet.cell(row=i, column=x).value, str):
                    # print('Колонка ' + str(x) + ' : ' + sheet.cell(row=i, column=x).value )
                    cmd = cmd + '\'' + sheet.cell(row=i, column=x).value + "', "
                # elif isinstance(sheet.cell(row=i, column=x).value, int):
                #     numb = sheet.cell(row=i, column=x).value
                #     cmd = cmd + '\'' + str(numb) + "', "
                elif isinstance(sheet.cell(row=i, column=x).value, datetime.datetime):
                    date_time = sheet.cell(row=i, column=x).value
                    date_time_str = date_time.strftime("%d.%m.%Y")
                    # print('Колонка ' + str(x) + ' : ' + date_time_str + ', in')
                    cmd = cmd + '\'' + date_time_str + "', "
                else:
                    cmd = cmd + '\'' + str(sheet.cell(row=i, column=x).value).replace('.', ',') + "', "
            cmd = cmd + "0 )"
            count_rows = count_rows+1
            print(str(i) + ' : ' + cmd)
            cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()
    now = datetime.datetime.now()
    print("Загружено записей: " + str(count_rows) + ' : ' + now.strftime("%d-%m-%Y %H:%M:%S"))
    return


def update_pm(table_name):
    print("Обновление PM стартовало. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    cmd = "DECLARE " \
          " cnt pls_integer default 0;" \
          "BEGIN " \
          "FOR rec IN (SELECT * from " + table_name + " m where m.pm is null) "\
          "LOOP "\
          "UPDATE " + table_name + " m " \
          "SET pm = ( " \
          "  SELECT SUM(CEIL(months_between(mp.stopdate, mp.d_naz))) " \
          "  FROM ss_m_sol ms, ss_m_pay mp " \
          "  WHERE mp.id = ms.mpay " \
          "  AND ms.sicid = m.sicid " \
          "  AND substr(mp.pc,1,4)='0703' " \
          "  AND mp.nsum > 0 " \
          "  AND mp.id = (SELECT MAX(ID) FROM ss_m_pay WHERE soliray_id = mp.soliray_id) " \
          "  ) " \
          "WHERE m.iin = rec.iin "\
          "AND m.mnth = rec.mnth; "\
          "if cnt>100 then cnt:=0; commit; end if; "\
          "cnt:=cnt+1; "\
          "END LOOP; "\
          "commit; "\
          "END;"
    print('CMD: ' + cmd)
    con = get_connection()
    cursor = con.cursor()
    cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()
    print("Обновление PM завершено. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))


def create_table_2(t_name):
    cmd = "CREATE TABLE " + t_name + "_2 as " \
    "SELECT p.iin, p.mnth, COUNT(DISTINCT CASE WHEN ip.pay_month <= ADD_MONTHS(p.reg_date, -1) THEN ip.pay_month ELSE NULL END) su, " \
    "SUM(CASE WHEN ip.pay_month BETWEEN trunc(ADD_MONTHS(p.reg_date, -24), 'MM') AND ADD_MONTHS(p.reg_date, -1) THEN s ELSE 0 END) so24, " \
    "round(SUM(CASE WHEN pay_month BETWEEN trunc(ADD_MONTHS(p.reg_date, -24), 'MM') AND ADD_MONTHS(p.reg_date, -1) THEN s * 100 / PERC ELSE 0 END), 2) sd, " \
    "COUNT(distinct CASE WHEN ip.pay_month BETWEEN ADD_MONTHS(p.reg_date, -24) AND ADD_MONTHS(p.reg_date, -1) THEN ip.pay_month ELSE NULL END) su24 " \
    "FROM " + t_name + " p, " \
    "(SELECT br.sicid, pay_month, P_RNN, np.perc, LEAST(SUM(SUM_PAY), BS.BASE_SIZE / 100 * NP.PERC * 10) S, " \
    " mnth , ip.knp " \
    " FROM si_member_2 IP, RFBS_BASE_SIZE BS, NNN_PERC np, " + t_name + " br WHERE BS.BASE_TYPE = 5 " \
    " AND br.sicid = ip.sicid " \
    " and (ip.type_payer!='Е' or ip.type_payer is null) " \
    " AND IP.PAY_MONTH BETWEEN BS.DATE_BEG AND NVL(BS.DATE_END, ADD_MONTHS(TRUNC(SYSDATE, 'YY'), 12) - 1) " \
    " AND ip.KNP in ('012') " \
    " AND IP.PAY_MONTH = NP.MNT " \
    " and not exists (select a.sicid from  v_gfss_incoming_pay_sl_with026 a where ip.sicid=a.sicid " \
    " and ip.pay_month=a.pay_month and a.sum_pay=ip.sum_pay ) " \
    " GROUP BY br.sicid, P_RNN, PAY_MONTH, BS.BASE_SIZE, np.perc, reg_date , mnth ,ip.knp " \
    " ) ip " \
    " WHERE p.sicid = ip.sicid " \
    "AND p.mnth = ip.mnth " \
    "GROUP BY p.iin, p.mnth"
    print(cmd)
    con = get_connection()
    cursor = con.cursor()
    try:
        print("Удаляем таблицу " + t_name + '_2')
        cursor.execute("drop table " + t_name + '_2')
        print("Удалена таблица " + t_name + '_2')
    except cx_Oracle.Error:
        print("Error import cx_Oracle :", cx_Oracle.DataError)

    cursor.execute(cmd)
    print("Вспомогательная таблица создана. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cursor.close()
    con.close()
    return


def get_result(t_name):
    cmd = ' SELECT ID "№ п/п", ' \
            'REGION "Код отделения", mb.cz_code "Код центра занятости", mb.cz_name "Наименование кода ЦЗ", ' \
            'mb.IIN "ИИН", mb.lastname "Фамилия", mb.firstname "Имя", mb.middlename "Отчество", '\
            'to_char(mb.birthday,\'dd.mm.yyyy\') "Дата рождения", '\
            'to_char(mb.reg_date,\'dd.mm.yyyy\') "Дата регистрации в ЦЗ", mb.status_code "Код статуса заявителя ЦЗ", ' \
            'mb.status_in_cz "Статуса заявителя ЦЗ", mb.member "(Не-)Участник системы СС", ' \
            'mb.kzd "КЗД", ' \
            'm3.su "общий стаж участия", ' \
            'CASE WHEN m3.so24 > 0 THEN \'Да\' ELSE \'Нет\' END so, ' \
            'm3.su24 "Кол-во мес за последние 24 мес", m3.so24, ' \
            'round(m3.sd / 24, 2) sd , ' \
            'CASE WHEN mb.pm > 0 THEN \'Да\' ELSE \'Нет\' END is0703, ' \
            'mb.pm , '\
            'm3.su - (12 * nvl(mb.pm, 0)) nsu , '\
            'CASE ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 0 AND 5 THEN 0.1 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 6 AND 11 THEN 0.7 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 12 AND 23 THEN 0.75 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 24 AND 35 THEN 0.85 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 36 AND 47 THEN 0.9 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 48 AND 59 THEN 0.95 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 60 AND 71 THEN 1 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 72 AND 83 THEN 1.02 ' \
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 84 AND 95 THEN 1.04 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 96 AND 107 THEN 1.06 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 108 AND 119 THEN 1.08 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 120 AND 131 THEN 1.1 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 132 AND 143 THEN 1.12 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 144 AND 155 THEN 1.14 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 156 AND 167 THEN 1.16 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 168 AND 179 THEN 1.18 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 180 AND 191 THEN 1.2 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 192 AND 203 THEN 1.22 '\
            'END ksu_new , '\
            'CASE '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 0 AND 5 THEN 0 '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 6 AND 11 THEN  ceil(0.70 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 12 AND 23 THEN ceil(0.75 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 24 AND 35 THEN ceil(0.85 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 36 AND 47 THEN ceil(0.90 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 48 AND 59 THEN ceil(0.95 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 60 AND 71 THEN ceil(1 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 72 AND 83 THEN ceil(1.02 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 84 AND 95 THEN ceil(1.04 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 96 AND 107 THEN ceil(1.06 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 108 AND 119 THEN ceil(1.08 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 120 AND 131 THEN ceil(1.1 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 132 AND 143 THEN ceil(1.12 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 144 AND 155 THEN ceil(1.14 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 156 AND 167 THEN ceil(1.16 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 168 AND 179 THEN ceil(1.18 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 180 AND 191 THEN ceil(1.2 * m3.sd / 24 * 0.4) '\
            'WHEN m3.su - (12 * nvl(mb.pm, 0)) BETWEEN 192 AND 203 THEN ceil(1.22 * m3.sd / 24 * 0.4) '\
            'END size_new '\
            'FROM ' + t_name + ' mb, ' + t_name + '_2 m3 ' \
            'WHERE mb.iin = m3.iin(+) '\
            'AND mb.mnth = m3.mnth(+) '\
            'ORDER BY ID'
    con = get_connection()
    cursor = con.cursor()
    print(cmd)
    cursor.execute(cmd)
    cursor.rowfactory = ResultF
    return cursor


def print_report(table_name):
    file_name = table_name + '.xlsx'
    file_path = cfg.REPORTS_PATH + file_name

    if os.path.isfile(file_path):
        print('Файл уже существует: ' + file_path)
        return file_name

    workbook = xlsxwriter.Workbook(file_path)
    worksheet = workbook.add_worksheet()

    cursor = get_result(table_name)

    # print("Провели расчет и формируем Excel: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    title1_cell_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'size': 18})
    # worksheet.merge_range('B1:H1', "", title1_cell_format)
    # Устанавливаем высоту первой строки
    worksheet.set_row(0, 72)

    title2_cell_format = workbook.add_format(
        {'align': 'left', 'bold': True, 'valign': 'vcenter', 'size': 14, 'underline': True})
    title3_cell_format = workbook.add_format({'align': 'left', 'bold': True, 'valign': 'vcenter', 'size': 12})
    title3_1_cell_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'size': 12})
    theme_name_format = workbook.add_format({'align': 'left', 'valign': 'vcenter', 'border': 1, 'text_wrap': True})

    common_format = workbook.add_format({'align': 'center', 'valign': 'vcenter', 'border': 1})
    common_format.set_text_wrap()
    common_format_2 = workbook.add_format({'align': 'center', 'bold': True, 'border': 1})
    common_format_2.set_align('vcenter')
    common_format_2.set_text_wrap()
    name_format = workbook.add_format({'font_color': 'black', 'align': 'left', 'valign': 'vcenter', 'border': 1})
    sum_pay_format = workbook.add_format({'num_format': '#,###,##0.00', 'font_color': 'black', 'valign': 'vcenter', 'border': 1})
    date_format = workbook.add_format({'num_format': 'dd/mm/yy', 'align': 'center', 'valign': 'vcenter', 'border': 1})


    worksheet.set_column(0, 0, 5)
    worksheet.write('A1', '№', common_format_2)
    worksheet.set_column(1, 1, 12)
    worksheet.write('B1', 'Код отделения', common_format_2)
    worksheet.set_column(2, 2, 14)
    worksheet.write('C1', 'Код центра занятости', common_format_2)
    worksheet.set_column(3, 3, 96)
    worksheet.write('D1', 'Наименование центра занятости', common_format_2)
    worksheet.set_column(4, 4, 13)
    worksheet.write('E1', 'ИИН', common_format_2)
    worksheet.set_column(5, 5, 24)
    worksheet.write('F1', 'Фамилия', common_format_2)
    worksheet.set_column(6, 6, 24)
    worksheet.write('G1', 'Имя', common_format_2)
    worksheet.set_column(7, 8, 24)
    worksheet.write('H1', 'Отчество', common_format_2)
    worksheet.set_column(8, 8, 12)   # Дата рождения
    worksheet.write('I1', 'Дата рождения', common_format_2)
    worksheet.set_column(9, 9, 12)   # Дата регистрации
    worksheet.write('J1', 'Дата регистрации', common_format_2)
    worksheet.set_column(10, 10, 7)
    worksheet.write('K1', 'Код статуса заявителя ЦЗ', common_format_2)
    worksheet.set_column(11, 11, 10)
    worksheet.write('L1', 'Статус заявителя ЦЗ', common_format_2)
    worksheet.set_column(12, 12, 14)
    worksheet.write('M1', '(Не-) Участник системы СС', common_format_2)
    worksheet.set_column(13, 13, 8)  # Общий стаж участия
    worksheet.write('N1', 'КЗД', common_format_2)
    worksheet.set_column(14, 14, 12)  # Общий стаж участия
    worksheet.write('O1', 'Общий стаж участия', common_format_2)
    worksheet.set_column(15, 15, 13)  # Кол-во СО за последние 24 месяца
    worksheet.write('P1', 'Есть СО за последние 24 месяца', common_format_2)
    worksheet.set_column(16, 16, 13)  # Кол-во СО за последние 24 месяца
    worksheet.write('Q1', 'Кол-во мес за последние 24 мес', common_format_2)
    worksheet.set_column(17, 17, 12)  # Кол-во СО за последние 24 месяца
    worksheet.write('R1', 'СО 24 мес', common_format_2)
    worksheet.set_column(18, 18, 14)  # Кол-во СО за последние 24 месяца
    worksheet.write('S1', 'СМД', common_format_2)
    worksheet.set_column(19, 19, 15)
    worksheet.write('T1', 'Вид выплаты по  СВпр (если есть предыдущее назначение ДА/НЕТ)', common_format_2)
    worksheet.set_column(20, 20, 15)
    worksheet.write('U1', 'Кол-во месяцев назначения с учетом решения на прекращение', common_format_2)
    worksheet.set_column(21, 21, 15)
    worksheet.write('V1', 'Стаж участия с учетом ранее назначенного периода', common_format_2)
    worksheet.set_column(22, 22, 10)
    worksheet.write('W1', 'КСУ', common_format_2)
    worksheet.set_column(23, 23, 12)
    worksheet.write('X1', 'Предполагаемый размер', common_format_2)



    row = 0
    first_record = 1
    for record in cursor:
        worksheet.write(row + first_record, 0, record.id, common_format)
        worksheet.write(row + first_record, 1, record.region, common_format)
        worksheet.write(row + first_record, 2, record.cz_code, common_format)
        worksheet.write(row + first_record, 3, record.cz_name, name_format)
        worksheet.write(row + first_record, 4, record.iin, common_format)
        worksheet.write(row + first_record, 5, record.lastname, name_format)
        worksheet.write(row + first_record, 6, record.firstname, name_format)
        worksheet.write(row + first_record, 7, record.middlename, name_format)
        worksheet.write(row + first_record, 8, record.birthday, date_format)
        worksheet.write(row + first_record, 9, record.reg_date, date_format)
        worksheet.write(row + first_record, 10, record.status_code, common_format)
        worksheet.write(row + first_record, 11, record.status_in_cz, common_format)
        worksheet.write(row + first_record, 12, record.member, common_format)
        worksheet.write(row + first_record, 13, record.kzd, common_format)
        worksheet.write(row + first_record, 14, record.su, common_format)
        worksheet.write(row + first_record, 15, record.so, common_format)
        worksheet.write(row + first_record, 16, record.su24, common_format)
        worksheet.write(row + first_record, 17, record.so24, sum_pay_format)
        worksheet.write(row + first_record, 18, record.sd, sum_pay_format)
        worksheet.write(row + first_record, 19, record.is0703, common_format)
        worksheet.write(row + first_record, 20, record.pm, common_format)
        worksheet.write(row + first_record, 21, record.nsu, common_format)
        worksheet.write(row + first_record, 22, record.ksu_new, common_format)
        worksheet.write(row + first_record, 23, record.size_new, sum_pay_format)
        row += 1

    print("Отчет сформирован. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    workbook.close()

    # print("Завершен расчет: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    if cfg.print_at_once:
        None
    #        my_printer = win32print.GetDefaultPrinter()
    #        win32api.ShellExecute(
    #            0,
    #            "print",
    #            file_name,
    #            my_printer,
    #            ".",
    #            0
    #        )
    return file_name


def update_iin(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Обновление ИИН для: " + table_name + ' : '+ datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'update ' + table_name + ' set iin=\'0\'||iin  where sicid is null and length(iin)<12'
    print(cmd)
    for i in range(1, 6):
        cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()


def update_region(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Обновление Region для: " + table_name + ' : '+ datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'update ' + table_name + ' set region=\'0\'||region  where length(region)<4'
    print(cmd)
    for i in range(1, 6):
        cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()


if __name__ == "__main__":
    print("Начало работы программы: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    t_name = create_table_name()
    # create_table(t_name)
    # load_quart_unemployed(t_name)
    # create_index(t_name)
    # update_iin(t_name)
    # update_region(t_name)
    # fill_sicid_mnth(t_name)
    # t_name_2 = create_table_2(t_name)
    # update_pm(t_name)
    print_report(t_name)