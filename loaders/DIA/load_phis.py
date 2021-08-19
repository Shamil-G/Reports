from db_oracle.connect import get_connection
import cx_Oracle
import configs.config_130721 as cfg
from openpyxl import load_workbook
import datetime
import os.path
import xlsxwriter


def create_table_name():
    table_name_1 = "TMP_DIA_2020" + "_13072021"
    table_name_2 = "TMP_DIA_2021" + "_13072021"
    return table_name_1, table_name_2


def create_table(table_name):
    con = get_connection()
    cursor = con.cursor()
    try:
        print("Удаляем таблицу " + table_name)
        cursor.execute("drop table " + table_name)
        print("Удалена таблица " + table_name)
    except cx_Oracle.Error:
        print("Error import cx_Oracle :", cx_Oracle.DataError)

    cmd = 'create table ' + table_name + ' '\
        '( sv_tp nvarchar2(12), sicid number(12), id NUMBER(8), region VARCHAR2(4), rfpm_id varchar2(8), iin varchar2(12), '\
        'fio nvarchar2(128), sex nchar(1), date_risk date, smd number(19,2), kzd number(6,2), mzp number(8,2), '\
        'cnt_mnth number(3), sum_to_pay number(19,2), status nchar(18), avg_sum number(19,2) )'
    cursor.execute(cmd)
    print("Создана таблица " + table_name)
    con.commit()
    cursor.close()
    con.close()


def get_file_name():
    f_name_1 = cfg.file_1
    f_name_2 = cfg.file_2
    return f_name_1, f_name_2

def load_table(table_name, f_name):
    # Нормируем путь к файлу по слэшам
    f_path = cfg.REPORTS_PATH
    path = f_path + '\\' + f_name
    file_path = os.path.normpath(path)
    print("Работаем с файлом: " + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return

    print("Загрузка стартовала: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    wb = load_workbook(file_path)
    print("Книга загружена: " + path)
    # sheet = wb.active
    # Создадим новое задание

    con = get_connection()
    cursor = con.cursor()

    # print('SHEET name :' + wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print('Загружается книга: ' + sheet_name + ' : ' + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
        print("Всего в книге строк: " + str(sheet.max_row))
        count_rows = 0
        for i in range(cfg.first_row, sheet.max_row+1):
            con.commit()
            if not sheet.cell(row=i, column=1).value:
                break
            cmd = 'insert into ' + table_name + ' ' \
               '( sv_tp, id, region, rfpm_id, iin, fio, sex, date_risk, smd, kzd, mzp, cnt_mnth, sum_to_pay, status) '\
                'values ( \'' + sheet_name + '\', '
            for x in range(1, 13):
                if isinstance(sheet.cell(row=i, column=x).value, str):
                    # print('Колонка ' + str(x) + ' : ' + sheet.cell(row=i, column=x).value )
                    cmd = cmd + '\'' + sheet.cell(row=i, column=x).value + "', "
                # elif isinstance(sheet.cell(row=i, column=x).value, int):
                #     numb = sheet.cell(row=i, column=x).value
                #     cmd = cmd + '\'' + str(numb) + "', "
                elif isinstance(sheet.cell(row=i, column=x).value, datetime.datetime):
                    date_time = sheet.cell(row=i, column=x).value
                    date_time_str = date_time.strftime("%d.%m.%Y")
                    # print('Колонка ' + str(x) + ' : ' + date_time_str + ', in')
                    cmd = cmd + '\'' + date_time_str + "', "
                else:
                    cmd = cmd + '\'' + str(sheet.cell(row=i, column=x).value).replace('.', ',') + "', "
            cmd = cmd + "0 )"
            count_rows = count_rows+1
            # Для тестирования раскомментарить
            # print(str(i) + ' : ' + cmd)
            cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()
    print("Загружено записей: " + str(count_rows) + ' : ' +  datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    return


def create_index(table_name):
    con = get_connection()
    cursor = con.cursor()
    cursor.execute('create index xn_' + table_name + 'iin on ' + table_name + ' (iin)')
    cursor.execute('create index xn_' + table_name + 'sicid on ' + table_name + ' (sicid)')
    cursor.close()
    con.close()


def update_iin(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Обновление ИИН для: " + table_name + ' : '+ datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'update ' + table_name + ' set iin=\'0\'||iin  where sicid is null and length(iin)<12'
    print(cmd)
    for i in range(1, 6):
        cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()


def fill_sicid(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Начато обновление SIСID. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'begin ' \
        'for cur in (select t.rowid, p.sicid, t.iin from ' + table_name + ' t, person p ' \
        'where t.sicid is null and p.rn=t.iin) ' \
        ' loop ' \
        '   update ' + table_name + ' t2 ' \
        '   set t2.sicid = cur.sicid ' \
        '   where t2.iin = cur.iin; ' \
        ' end loop; commit; end;'
    print('CMD: ' + cmd)
    cursor.execute(cmd)


def set_status(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Начато обновление STATUS. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = ''


if __name__ == "__main__":
    print("Начало работы программы: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    t_name_1, t_name_2 = create_table_name()
    file_1, file_2 = get_file_name()
    print('Таблицы к загрузке: ' + t_name_1 + ', ' + t_name_2)
    # print('Файлы к загрузке: ' + file_1 + ', ' + file_2)
    # create_table(t_name_1)
    # create_table(t_name_2)
    # load_table(t_name_1, file_1)
    # load_table(t_name_2, file_2)
    # create_index(t_name_1)
    # create_index(t_name_2)
    update_iin(t_name_1)
    update_iin(t_name_2)
    fill_sicid(t_name_1)
    fill_sicid(t_name_2)
    # set_status(t_name)
    # update_pm(t_name)
    # t_name_2 = create_table_2(t_name)
    # print_report(t_name)