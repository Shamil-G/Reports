from   connect import get_connection
import cx_Oracle
import config_dasorp_26_08_21 as cfg
from openpyxl import load_workbook
import datetime
import os.path
import xlsxwriter


def create_table(table_name):
    con = get_connection()
    cursor = con.cursor()
    try:
        print("Удаляем таблицу " + table_name)
        cursor.execute("drop table " + table_name)
        print("Удалена таблица " + table_name)
    except cx_Oracle.Error:
        print("Error import cx_Oracle :", cx_Oracle.DataError)

    cmd = 'create table ' + table_name + ' '\
        '( num nvarchar2(6), id number(6), name_region nvarchar2(128), kod_ugd varchar2(4), org_name nvarchar2(128), name_plat nvarchar2(256), '\
        'iin varchar2(12), rnn nvarchar2(12),  fio_ruk nvarchar2(128), sum_debt number(19,2), status nvarchar2(128), last_so_date date, last_so_sum number(19,2) )'
    print('cmd: ' + cmd)
    cursor.execute(cmd)
    print("Создана таблица " + table_name)
    con.commit()
    cursor.close()
    con.close()


# def get_last_so(iin):
#   print('GET Last SO. iin: ' + iin)
#   cmd = 'select * from ( '\
#	 'select /* first_rows(1) */ '\
#         'dl.pay_date, dl.pay_sum, dl.sicid, p.rn '\
#         'from   person p, pmdl_doc_list dl '\
#         'where  p.rn = \'' + iin + '\' '\ 
#         'and    p.sicid=dl.sicid '\
#         'order  by pay_date desc '\
#         ') where rownum=1'
#    print('CMD: '+cmd)
#    con = get_connection()
#    cursor = con.cursor()


def load_table(table_name, f_name):
    # Нормируем путь к файлу по слэшам
    f_path = cfg.REPORTS_PATH
    path = f_path + '\\' + f_name
    file_path = os.path.normpath(path)
    print("Работаем с файлом: " + file_path)

    if not os.path.isfile(file_path):
        print("File not exists: " + str(os.path.isfile(file_path)))
        return

    print("Загрузка стартовала: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))

    wb = load_workbook(file_path)
    print("Книга загружена: " + path)
    # sheet = wb.active
    # Создадим новое задание

    con = get_connection()
    cursor = con.cursor()

    # print('SHEET name :' + wb.sheetnames)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print('Загружается книга: ' + sheet_name + ' : ' + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
        print("Всего в книге строк: " + str(sheet.max_row))
        count_rows = 0
        for i in range(cfg.first_row, sheet.max_row+1):
            con.commit()
            if not sheet.cell(row=i, column=1).value:
                break
            cmd = 'insert into ' + table_name + ' ' \
               '( num, id, name_region, kod_ugd, org_name, name_plat, '\
               'iin, rnn,  fio_ruk, sum_debt, status, last_so_sum ) '\
                'values ( '
            for x in range(1, 12):
                if isinstance(sheet.cell(row=i, column=x).value, str):
                    # print('Колонка ' + str(x) + ' : ' + sheet.cell(row=i, column=x).value )
                    cmd = cmd + '\'' + sheet.cell(row=i, column=x).value + "', "
                # elif isinstance(sheet.cell(row=i, column=x).value, int):
                #     numb = sheet.cell(row=i, column=x).value
                #     cmd = cmd + '\'' + str(numb) + "', "
                elif isinstance(sheet.cell(row=i, column=x).value, datetime.datetime):
                    date_time = sheet.cell(row=i, column=x).value
                    date_time_str = date_time.strftime("%d.%m.%Y")
                    # print('Колонка ' + str(x) + ' : ' + date_time_str + ', in')
                    cmd = cmd + '\'' + date_time_str + "', "
                else:
                    cmd = cmd + '\'' + str(sheet.cell(row=i, column=x).value).replace('.', ',') + "', "
            cmd = cmd + "0 )"
            count_rows = count_rows+1
            # Для тестирования раскомментарить
            # print(str(i) + ' : ' + cmd)
            cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()
    print("Загружено записей: " + str(count_rows) + ' : ' +  datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    return


def create_index(table_name):
    con = get_connection()
    cursor = con.cursor()
    cursor.execute('create index xn_' + table_name + 'iin on ' + table_name + ' (iin)')
    cursor.close()
    con.close()


def update_iin(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Обновление ИИН для: " + table_name + ' : '+ datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'update ' + table_name + ' set iin=\'0\'||iin  where sicid is null and length(iin)<12'
    print(cmd)
    for i in range(1, 6):
        cursor.execute(cmd)
    con.commit()
    cursor.close()
    con.close()


def fill_sicid(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Начато обновление SIСID. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = 'begin ' \
        'for cur in (select t.rowid, p.sicid, t.iin from ' + table_name + ' t, person p ' \
        'where t.sicid is null and p.rn=t.iin) ' \
        ' loop ' \
        '   update ' + table_name + ' t2 ' \
        '   set t2.sicid = cur.sicid ' \
        '   where t2.iin = cur.iin; ' \
        ' end loop; commit; end;'
    print('CMD: ' + cmd)
    cursor.execute(cmd)


def set_status(table_name):
    con = get_connection()
    cursor = con.cursor()
    print("Начато обновление STATUS. " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    cmd = ''


if __name__ == "__main__":
    print("Начало работы программы: " + datetime.datetime.now().strftime("%d-%m-%Y %H:%M:%S"))
    print('Файл к загрузке: ' + cfg.file)
    # create_table(cfg.table_name)
    # load_table(cfg.table_name, cfg.file)
    create_index(cfg.table_name)
    # create_index(t_name_2)
    # set_status(t_name)
    # update_pm(t_name)
    # t_name_2 = create_table_2(t_name)
    # print_report(t_name)
